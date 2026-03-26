"""
services.py — All business logic for the examination module.

Functions here orchestrate data via selectors, perform calculations,
generate files (CSV / Excel / PDF), and return plain data or file objects.
No HTTP request/response handling lives here.
"""

import csv
import json
import os
import traceback
from collections import OrderedDict, defaultdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO, StringIO

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction, connection
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from applications.academic_information.models import Student
from applications.academic_procedures.models import course_registration
from applications.online_cms.models import Student_grades
from applications.programme_curriculum.models import (
    Course as Courses,
    CourseInstructor,
)

from . import selectors
from .models import (
    ALLOWED_GRADES,
    GRADE_CONVERSION,
    PBI_AND_BTP_ALLOWED_GRADES,
    PBI_BTP_COURSE_CODES,
    ALL_DISPLAY_GRADES,
    UG_PROGRAMMES,
    PG_PROGRAMMES,
    hidden_grades,
    ResultAnnouncement,
)

User = get_user_model()


# ---------------------------------------------------------------------------
# Utility / calculation helpers
# ---------------------------------------------------------------------------

def parse_academic_year(academic_year, semester_type):
    """
    Parse academic_year string (e.g., "2024-25") and return (working_year, session).
    """
    parts = academic_year.split("-")
    if len(parts) != 2:
        raise ValueError("Invalid academic year format. Expected format like '2024-25'.")
    first_year = parts[0].strip()
    second_year = parts[1].strip()
    if semester_type == "Odd Semester":
        working_year = int(first_year)
    else:
        working_year = int("20" + second_year)
    session = academic_year
    return working_year, session


def round_from_last_decimal(number, decimal_places=1):
    d = Decimal(str(number))
    return Decimal(d).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)


def is_valid_grade(grade_str, course_code):
    """Return True if the grade is valid for the given course code."""
    if not grade_str or not course_code:
        return False
    code = course_code.strip().upper()
    g = grade_str.strip().upper()
    if code in PBI_BTP_COURSE_CODES:
        return g in PBI_AND_BTP_ALLOWED_GRADES
    return g in ALLOWED_GRADES


def get_programme_list_or_error(programme_type):
    """Return programme list or raise ValueError for invalid type."""
    if not programme_type:
        return None
    key = programme_type.upper()
    if key == "UG":
        return list(UG_PROGRAMMES)
    if key == "PG":
        return list(PG_PROGRAMMES)
    raise ValueError("Invalid programme_type. Must be 'UG' or 'PG'.")


def format_semester_display(semester_no, semester_type=None, semester_label=None):
    """Format semester number for display in PDFs."""
    if semester_label and "summer" in semester_label.lower():
        return semester_label
    if semester_type and "summer" in semester_type.lower():
        if semester_no == 2:
            return "Summer 1"
        elif semester_no == 4:
            return "Summer 2"
        elif semester_no == 6:
            return "Summer 3"
        elif semester_no == 8:
            return "Summer 4"
        else:
            return f"Summer {semester_no // 2}"
    return str(semester_no)


def make_semester_label(no, sem_type):
    """Create label like 'Semester 3' or 'Summer 2'."""
    if no % 2 == 1:
        return f"Semester {no}"
    if sem_type == "Summer Semester":
        return f"Summer {no // 2}"
    return f"Semester {no}"


def trace_registration(reg_id, mapping):
    """Follow replacement chain to find original registration."""
    seen = set()
    while reg_id in mapping and reg_id not in seen:
        seen.add(reg_id)
        reg_id = mapping[reg_id]
    return reg_id


def gather_related_registrations(initial_reg, max_semester):
    """BFS to collect all related registrations via replacements."""
    from applications.academic_procedures.models import course_replacement

    related = set()
    queue = [initial_reg]
    while queue:
        reg = queue.pop(0)
        if reg.id in related:
            continue
        related.add(reg.id)
        olds = course_replacement.objects.filter(old_course_registration=reg)
        news = course_replacement.objects.filter(new_course_registration=reg)
        for rep in list(olds) + list(news):
            for neighbor in (rep.old_course_registration, rep.new_course_registration):
                if (
                    neighbor.student_id == initial_reg.student_id
                    and neighbor.semester_id.semester_no <= max_semester
                ):
                    queue.append(neighbor)
    return course_registration.objects.filter(id__in=related).exclude(
        id=initial_reg.id
    )


# ---------------------------------------------------------------------------
# SPI / CPI calculations
# ---------------------------------------------------------------------------

def calculate_spi_for_student(student, selected_semester, semester_type):
    semester_unit = Decimal("0")
    grades = selectors.get_grades_for_student_semester(
        student.id_id, selected_semester, semester_type
    )
    total_points = Decimal("0")
    total_credits = Decimal("0")
    for g in grades:
        credit = Decimal(str(g.course_id.credit))
        factor = GRADE_CONVERSION.get(g.grade.strip(), -1)
        if factor >= 0:
            if factor != 0:
                factor = Decimal(str(factor))
                total_points += factor * credit
                total_credits += credit
            semester_unit += credit
    spi = (
        round_from_last_decimal(Decimal("10") * (total_points / total_credits))
        if total_credits
        else 0
    )
    return spi, semester_unit, (total_points * 10)


def calculate_cpi_for_student(student, selected_semester, semester_type):
    total_unit = Decimal("0")
    grades = selectors.get_grades_up_to_semester(
        student.id_id, selected_semester, semester_type
    )
    registrations = selectors.get_registrations_up_to_semester(
        student, selected_semester, semester_type
    )

    reg_mapping = {}
    for reg in registrations:
        key = (reg.course_id.code.strip(), reg.semester_id.semester_no, reg.semester_type)
        reg_mapping[key] = reg.id

    replacements = selectors.get_course_replacements_for_student(student)
    reg_replacement_map = {}
    for rep in replacements:
        old_reg_id = rep.old_course_registration.id
        new_reg_id = rep.new_course_registration.id
        if new_reg_id != old_reg_id:
            reg_replacement_map[new_reg_id] = old_reg_id

    grade_groups = defaultdict(list)
    for g in grades:
        key = (g.course_id.code.strip(), g.semester, g.semester_type)
        reg_id = reg_mapping.get(key)
        if reg_id is None:
            continue
        original_reg_id = trace_registration(reg_id, reg_replacement_map)
        grade_groups[original_reg_id].append(g)

    total_points = Decimal("0")
    total_credits = Decimal("0")
    for orig_reg, g_list in grade_groups.items():
        best_record = max(
            g_list, key=lambda r: GRADE_CONVERSION.get(r.grade.strip(), -1)
        )
        grade_factor = GRADE_CONVERSION.get(best_record.grade.strip(), -1)
        credit = Decimal(str(getattr(best_record.course_id, "credit", 3)))
        if grade_factor >= 0:
            if grade_factor != 0:
                grade_factor = Decimal(str(grade_factor))
                total_points += grade_factor * credit
                total_credits += credit
            total_unit += credit

    cpi = (
        round_from_last_decimal(Decimal("10") * (total_points / total_credits))
        if total_credits
        else 0
    )
    return cpi, total_unit, (total_points * 10)


# ---------------------------------------------------------------------------
# Grade points helper
# ---------------------------------------------------------------------------

def compute_grade_points(grade_str):
    """Return grade points (Decimal) for display."""
    raw = GRADE_CONVERSION.get(grade_str, 0)
    return Decimal(str(raw * 10)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)


# ---------------------------------------------------------------------------
# Role redirect
# ---------------------------------------------------------------------------

def get_redirect_url_for_role(role):
    role_map = {
        "Associate Professor": "/examination/submitGradesProf/",
        "Professor": "/examination/submitGradesProf/",
        "Assistant Professor": "/examination/submitGradesProf/",
        "acadadmin": "/examination/updateGrades/",
        "Dean Academic": "/examination/verifyGradesDean/",
    }
    return role_map.get(role, "/dashboard/")


# ---------------------------------------------------------------------------
# Download template CSV
# ---------------------------------------------------------------------------

def generate_download_template(course_id, session_year, semester_type, programme_type=None):
    """Generate CSV HttpResponse for grade entry template."""
    regs = selectors.get_course_registrations(
        course_id, session_year, semester_type, programme_type
    ).order_by("student_id_id")

    if not regs.exists():
        return None, _no_students_message(programme_type)

    course_obj = regs.first().course_id
    response = HttpResponse(content_type="text/csv")
    course_name_clean = course_obj.name.replace(" ", "_").replace("/", "-")[:50]
    semester_type_clean = semester_type.replace(" ", "_")
    filename = f"{course_name_clean}_{semester_type_clean}_{session_year}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(["roll_no", "name", "branch", "grade", "remarks", "semester"])

    for entry in regs:
        student_entry = entry.student_id
        student_user = User.objects.get(username=student_entry.id_id)
        branch_acronym = ""
        if student_entry.batch_id:
            try:
                branch_acronym = student_entry.batch_id.discipline.acronym
            except AttributeError:
                pass
        if not branch_acronym and student_entry.id.department:
            branch_acronym = student_entry.id.department.name
        semester_no = ""
        if entry.course_slot_id and entry.course_slot_id.semester:
            semester_no = entry.course_slot_id.semester.semester_no
        writer.writerow([
            student_entry.id_id,
            f"{student_user.first_name} {student_user.last_name}",
            branch_acronym,
            "",
            "",
            semester_no,
        ])

    return response, None


def _no_students_message(programme_type):
    if programme_type:
        name = "Undergraduate" if programme_type.upper() == "UG" else "Postgraduate"
        return f"No {name} students found in this course for the selected academic year and semester."
    return "No registration data found for the provided course, academic year, and semester type."


# ---------------------------------------------------------------------------
# Check course students
# ---------------------------------------------------------------------------

def check_course_students(course_id, session_year, semester_type, programme_type=None):
    regs = selectors.get_course_registrations(
        course_id, session_year, semester_type, programme_type
    )
    has_students = regs.exists()
    student_count = regs.count() if has_students else 0
    return has_students, student_count


# ---------------------------------------------------------------------------
# Upload grades (acadadmin)
# ---------------------------------------------------------------------------

def upload_grades_admin(course_id, academic_year, semester_type, csv_file):
    """Process CSV and upload grades for acadadmin. Returns (success_msg, error_msg)."""
    working_year, session = parse_academic_year(academic_year, semester_type)
    courses_info = selectors.get_course_by_id(course_id)

    regs = selectors.get_course_registrations(course_id, academic_year, semester_type)
    if not regs.exists():
        return None, "NO STUDENTS REGISTERED IN THIS COURSE FOR THE SELECTED SEMESTER."

    existing = selectors.get_student_grades(course_id, academic_year, semester_type)
    if existing.exists() and not existing.first().reSubmit:
        return None, "THIS COURSE HAS ALREADY BEEN SUBMITTED."

    decoded_file = csv_file.read().decode("utf-8").splitlines()
    reader = csv.DictReader(decoded_file)
    required_columns = ["roll_no", "grade", "remarks"]
    if not all(col in (reader.fieldnames or []) for col in required_columns):
        return None, "CSV file must contain the following columns: roll_no, grade, remarks."

    errors = []
    allowed_list = ", ".join(sorted(ALLOWED_GRADES))

    with transaction.atomic():
        for index, row in enumerate(reader, start=1):
            roll_no = row.get("roll_no")
            grade_val = row.get("grade")
            remarks = row.get("remarks", "")
            semester = row.get("semester", None)

            try:
                stud = selectors.get_student_by_roll(roll_no)
            except Student.DoesNotExist:
                errors.append(f"Row {index}: Student with roll_no {roll_no} does not exist.")
                continue

            if not selectors.get_student_registration_for_course(
                stud, courses_info, semester_type, academic_year
            ).exists():
                errors.append(
                    f"Row {index}: Student with roll_no {roll_no} is not registered for this course in the selected semester."
                )
                continue

            if not is_valid_grade(grade_val, courses_info.code):
                errors.append(
                    f"Row {index}: Invalid grade '{grade_val}' for roll_no {roll_no}. "
                    f"Allowed grades are: {allowed_list}."
                )
                continue

            semester = semester or stud.curr_semester_no
            batch = stud.batch

            try:
                Student_grades.objects.update_or_create(
                    roll_no=roll_no,
                    course_id_id=course_id,
                    year=working_year,
                    semester=semester,
                    batch=batch,
                    academic_year=academic_year,
                    semester_type=semester_type,
                    defaults={
                        "grade": grade_val,
                        "remarks": remarks,
                        "reSubmit": False,
                        "academic_year": session,
                        "semester_type": semester_type,
                    },
                )
            except Exception as create_err:
                errors.append(
                    f"Row {index}: Error creating/updating grade for student with roll_no {roll_no} - {str(create_err)}"
                )
                continue

        if errors:
            error_summary = "\n".join(f"- {msg}" for msg in errors)
            raise Exception(error_summary)

    return "Grades uploaded successfully.", None


# ---------------------------------------------------------------------------
# Upload grades (professor)
# ---------------------------------------------------------------------------

def upload_grades_prof(
    username, course_id, academic_year, semester_type, csv_file, programme_type=None
):
    """Process CSV and upload grades for professor. Returns (success_msg, error_msg)."""
    working_year, session = parse_academic_year(academic_year, semester_type)

    course = selectors.get_course_by_id(course_id)

    regs = selectors.get_course_registrations_by_working_year(
        course, working_year, semester_type
    )
    if not regs.exists():
        return None, "NO STUDENTS REGISTERED IN THIS COURSE FOR THE SELECTED SEMESTER."

    ug_student_ids = Student.objects.filter(programme__in=UG_PROGRAMMES).values_list("id", flat=True)
    pg_student_ids = Student.objects.filter(programme__in=PG_PROGRAMMES).values_list("id", flat=True)
    course_has_ug = regs.filter(student_id__in=ug_student_ids).exists()
    course_has_pg = regs.filter(student_id__in=pg_student_ids).exists()

    if programme_type:
        if programme_type.upper() == "UG":
            if not course_has_ug:
                return None, "No UG students registered in this course."
            regs = regs.filter(student_id__in=ug_student_ids)
        elif programme_type.upper() == "PG":
            if not course_has_pg:
                return None, "No PG students registered in this course."
            regs = regs.filter(student_id__in=pg_student_ids)
        else:
            return None, "Invalid programme_type. Must be 'UG' or 'PG'."
    else:
        if course_has_ug and course_has_pg:
            return None, json.dumps({
                "error": "This course has both UG and PG students. Please specify programme_type as 'UG' or 'PG'.",
                "course_info": {
                    "course_code": course.code,
                    "course_name": course.name,
                    "has_ug": course_has_ug,
                    "has_pg": course_has_pg,
                    "total_registrations": regs.count(),
                },
            })
        elif course_has_ug:
            programme_type = "UG"
            regs = regs.filter(student_id__in=ug_student_ids)
        elif course_has_pg:
            programme_type = "PG"
            regs = regs.filter(student_id__in=pg_student_ids)

    # Check existing submissions
    existing_query = Student_grades.objects.filter(
        course_id=course_id, academic_year=academic_year, semester_type=semester_type
    )
    if programme_type:
        student_rolls = [reg.student_id_id for reg in regs]
        existing_for_programme = existing_query.filter(roll_no__in=student_rolls)
        if existing_for_programme.exists():
            non_resubmit = existing_for_programme.filter(reSubmit=False)
            if non_resubmit.exists():
                label = "UG" if programme_type.upper() == "UG" else "PG"
                return None, f"THIS COURSE HAS ALREADY BEEN SUBMITTED FOR {label} STUDENTS."
    else:
        existing = existing_query.first()
        if existing and not existing.reSubmit:
            return None, "THIS COURSE HAS ALREADY BEEN SUBMITTED."

    # Instructor ownership
    if not selectors.check_instructor_ownership(course_id, username, working_year):
        return None, "ACCESS_DENIED:you are not assigned as instructor for this course."

    # Parse CSV
    decoded = csv_file.read().decode("utf-8").splitlines()
    reader = csv.DictReader(decoded)
    required_cols = {"roll_no", "grade", "remarks"}
    if not required_cols.issubset(reader.fieldnames or []):
        return None, "CSV file must contain columns: roll_no, grade, remarks."

    errors = []
    with transaction.atomic():
        # Reset reSubmit flags
        reset_query = Student_grades.objects.filter(
            course_id_id=course_id,
            academic_year=academic_year,
            semester_type=semester_type,
        )
        if programme_type:
            prog_rolls = [reg.student_id_id for reg in regs]
            reset_query = reset_query.filter(roll_no__in=prog_rolls)
        reset_query.update(reSubmit=False)

        for idx, row in enumerate(reader, start=1):
            roll_no = row.get("roll_no", "").strip()
            grade_val = row.get("grade", "").strip()
            remarks = row.get("remarks", "").strip()
            sem_csv = row.get("semester", "").strip() or None

            try:
                stud = selectors.get_student_by_roll(roll_no)
            except Student.DoesNotExist:
                errors.append(f"Row {idx}: Student with roll_no {roll_no} does not exist.")
                continue

            if not selectors.get_student_registration_for_course(
                stud, course, semester_type, academic_year
            ).exists():
                errors.append(
                    f"Row {idx}: Student {roll_no} not registered for this course/semester."
                )
                continue

            if programme_type:
                student_programme = stud.programme
                if programme_type.upper() == "UG" and student_programme not in UG_PROGRAMMES:
                    errors.append(
                        f"Row {idx}: Student {roll_no} is not a UG student (programme: {student_programme})."
                    )
                    continue
                elif programme_type.upper() == "PG" and student_programme not in PG_PROGRAMMES:
                    errors.append(
                        f"Row {idx}: Student {roll_no} is not a PG student (programme: {student_programme})."
                    )
                    continue

            if grade_val not in ALLOWED_GRADES:
                allowed = ", ".join(sorted(ALLOWED_GRADES))
                errors.append(f"Row {idx}: Invalid grade '{grade_val}'. Allowed: {allowed}.")
                continue

            semester = sem_csv or stud.curr_semester_no
            batch = stud.batch

            try:
                Student_grades.objects.update_or_create(
                    roll_no=roll_no,
                    course_id_id=course_id,
                    year=working_year,
                    batch=batch,
                    academic_year=academic_year,
                    semester_type=semester_type,
                    semester=semester,
                    defaults={
                        "grade": grade_val,
                        "remarks": remarks,
                        "reSubmit": False,
                    },
                )
            except Exception as exc:
                errors.append(f"Row {idx}: Error saving grade for {roll_no}: {str(exc)}")
                continue

        if errors:
            summary = "\n".join(f"- {e}" for e in errors)
            raise Exception(f"Upload failed with the following errors:\n{summary}")

    programme_msg = f" for {programme_type.upper()} students" if programme_type else ""
    return f"Grades uploaded successfully{programme_msg}.", None


# ---------------------------------------------------------------------------
# Moderate student grades
# ---------------------------------------------------------------------------

def moderate_student_grades(student_ids, semester_ids, course_ids, grades, remarks, allow_resubmission):
    """Update or create grade records. Returns CSV HttpResponse."""
    for student_id, semester_id, course_id, grade_val, remark in zip(
        student_ids, semester_ids, course_ids, grades, remarks
    ):
        try:
            grade_of_student = Student_grades.objects.get(
                course_id=course_id, roll_no=student_id, semester=semester_id
            )
            grade_of_student.remarks = remark
            grade_of_student.grade = grade_val
            grade_of_student.verified = True
            if allow_resubmission.upper() == "YES":
                grade_of_student.reSubmit = True
            grade_of_student.save()
        except Student_grades.DoesNotExist:
            hidden_grades.objects.create(
                course_id=course_id,
                student_id=student_id,
                semester_id=semester_id,
                grade=grade_val,
            )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="grades.csv"'
    writer = csv.writer(response)
    writer.writerow(["Student ID", "Semester ID", "Course ID", "Grade"])
    for student_id, semester_id, course_id, grade_val in zip(
        student_ids, semester_ids, course_ids, grades
    ):
        writer.writerow([student_id, semester_id, course_id, grade_val])
    return response


# ---------------------------------------------------------------------------
# Generate transcript data
# ---------------------------------------------------------------------------

def generate_transcript_data(student_id, semester_json):
    """Return dict with transcript data for a student/semester."""
    semester = json.loads(semester_json)
    semester_number = semester.get("no")
    semester_type = semester.get("type")

    student = selectors.get_student_by_roll(student_id)
    cpi, tu, _ = calculate_cpi_for_student(student, semester_number, semester_type)
    spi, su, _ = calculate_spi_for_student(student, semester_number, semester_type)

    courses_registered = selectors.get_student_grades_for_student(
        student_id, semester_number, semester_type
    )

    academic_year = None
    if courses_registered.exists():
        academic_year = courses_registered.first().academic_year

    course_grades = {}
    for reg in courses_registered:
        course = reg.course_id
        course_grades[course.id] = {
            "course_name": course.name,
            "course_code": course.code,
            "credit": course.credit,
            "grade": reg.grade,
            "points": compute_grade_points(reg.grade),
        }

    student_info = _build_student_info(student, academic_year)

    return {
        "name": student_info["name"],
        "student_name": student_info["name"],
        "roll_number": student_info["roll_number"],
        "programme": student_info["programme"],
        "department": student_info["department"],
        "branch": student_info["department"],
        "academic_year": academic_year or "",
        "student_info": student_info,
        "courses_grades": course_grades,
        "spi": spi,
        "cpi": cpi,
        "tu": tu,
        "su": su,
    }


# ---------------------------------------------------------------------------
# Student info builder
# ---------------------------------------------------------------------------

def _build_student_info(student, academic_year=None):
    return {
        "name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
        "student_name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
        "rollNumber": student.id.user.username,
        "roll_number": student.id.user.username,
        "programme": student.programme,
        "batch": str(student.batch_id) if student.batch_id else str(student.batch),
        "branch": student.id.department.name if student.id.department else "",
        "department": student.id.department.name if student.id.department else "",
        "semester": student.curr_semester_no,
        "academicYear": academic_year or "",
        "academic_year": academic_year or "",
    }


# ---------------------------------------------------------------------------
# Transcript form data
# ---------------------------------------------------------------------------

def get_transcript_form_data():
    """Return programmes, batches, specializations for the form."""
    batches_queryset = selectors.get_running_batches()
    batch_list = [
        {"id": batch.id, "label": f"{batch.name} - {batch.discipline} {batch.year}"}
        for batch in batches_queryset
    ]
    programmes = list(selectors.get_student_programmes())
    specializations = list(selectors.get_student_specializations())
    return {
        "programmes": programmes,
        "batches": batch_list,
        "specializations": specializations,
    }


def get_students_for_transcript(batch, specialization, semester):
    students = selectors.get_students_by_batch(batch, specialization)
    return {"students": list(students.values()), "semester": semester}


# ---------------------------------------------------------------------------
# Generate result Excel
# ---------------------------------------------------------------------------

def generate_result_excel(semester, batch_id, semester_type, branch=None):
    """Generate Excel HttpResponse with student grades, SPI, CPI."""
    batch_obj = selectors.get_batch_by_id(batch_id)
    if not batch_obj:
        return None, "Batch not found."

    students = selectors.get_students_by_batch(batch_id, branch)

    course_ids = (
        Student_grades.objects.filter(
            batch=batch_obj.year,
            semester=semester,
            roll_no__in=students,
            semester_type=semester_type,
        )
        .exclude(grade__isnull=True)
        .exclude(grade="")
        .values_list("course_id_id", flat=True)
        .distinct()
    )
    courses = Courses.objects.filter(id__in=course_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Grades"

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws["A1"] = "S. No"
    ws["B1"] = "Roll No"
    ws["C1"] = "Name"
    for col in ("A", "B", "C"):
        cell = ws[col + "1"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 18
    ws.column_dimensions[get_column_letter(3)].width = 30

    col_idx = 4
    for course in courses:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
        cell = ws.cell(row=1, column=col_idx)
        cell.value = course.code
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = header_fill

        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
        cell = ws.cell(row=2, column=col_idx)
        cell.value = course.name
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = header_fill

        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx + 1)
        cell = ws.cell(row=3, column=col_idx)
        cell.value = course.credit
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = header_fill

        for label, offset in [("Grade", 0), ("Remarks", 1)]:
            cell = ws.cell(row=4, column=col_idx + offset)
            cell.value = label
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

        ws.column_dimensions[get_column_letter(col_idx)].width = 25
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 25
        col_idx += 2

    for label in ["SPI", "CPI", "SU", "TU", "SP", "TP"]:
        cell = ws.cell(row=1, column=col_idx)
        cell.value = label
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = header_fill
        col_idx += 1

    max_col = ws.max_column
    for row in range(1, 5):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.border = thin_border

    row_idx = 5
    for idx, student in enumerate(students, start=1):
        ws.cell(row=row_idx, column=1).value = idx
        ws.cell(row=row_idx, column=2).value = student.id_id

        try:
            student_user = User.objects.get(username=student.id_id)
            student_name = (
                f"{student_user.first_name} {student_user.last_name}".strip()
                or student_user.username
            )
        except Exception:
            student_name = student.id_id

        ws.cell(row=row_idx, column=3).value = student_name
        ws.cell(row=row_idx, column=3).alignment = Alignment(
            horizontal="left", vertical="center"
        )

        student_grades = Student_grades.objects.filter(
            roll_no=student.id_id,
            course_id_id__in=course_ids,
            semester_type=semester_type,
            semester=semester,
        )
        grades_map = {g.course_id_id: g for g in student_grades}

        col_ptr = 4
        for course in courses:
            grade_entry = grades_map.get(course.id)
            grade_val = grade_entry.grade if grade_entry else "-"

            remark = "-"
            if grade_entry:
                reg = course_registration.objects.filter(
                    student_id=student,
                    course_id=course,
                    semester_id__semester_no=semester,
                    semester_type=semester_type,
                    session=grade_entry.academic_year,
                ).first()
                if reg:
                    related_regs = gather_related_registrations(reg, semester)
                    attempts = []
                    for r in related_regs:
                        g = Student_grades.objects.filter(
                            roll_no=student.id_id,
                            course_id__code=r.course_id.code,
                            semester=r.semester_id.semester_no,
                            semester_type=r.semester_type,
                            academic_year=r.session,
                        ).order_by("-semester").first()
                        if g:
                            attempts.append((r.course_id.code, g.grade))

                    if len(attempts) >= 1:
                        scored = sorted(
                            attempts,
                            key=lambda x: GRADE_CONVERSION.get(x[1], -1),
                            reverse=True,
                        )
                        first_code, first_grade = scored[0]
                        if first_grade in ("F", "X"):
                            remark = "R(BL)" if first_code == course.code else "S(BL)"
                        else:
                            remark = "R(IM)" if first_code == course.code else "S(IM)"

            ws.cell(row=row_idx, column=col_ptr).value = grade_val
            ws.cell(row=row_idx, column=col_ptr + 1).value = remark
            for c in [col_ptr, col_ptr + 1]:
                ws.cell(row=row_idx, column=c).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            col_ptr += 2

        spi_val, SU, SP = calculate_spi_for_student(student, semester, semester_type)
        cpi_val, TU, TP = calculate_cpi_for_student(student, semester, semester_type)
        ws.cell(row=row_idx, column=col_ptr).value = spi_val
        ws.cell(row=row_idx, column=col_ptr + 1).value = cpi_val
        ws.cell(row=row_idx, column=col_ptr + 2).value = SU
        ws.cell(row=row_idx, column=col_ptr + 3).value = TU
        ws.cell(row=row_idx, column=col_ptr + 4).value = SP
        ws.cell(row=row_idx, column=col_ptr + 5).value = TP
        for c in [col_ptr, col_ptr + 1]:
            ws.cell(row=row_idx, column=c).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        row_idx += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="student_grades.xlsx"'
    wb.save(response)
    return response, None


# ---------------------------------------------------------------------------
# Download excel CSV
# ---------------------------------------------------------------------------

def generate_download_excel_csv(student_ids, semester_ids, course_ids, grades):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="grades.csv"'
    writer = csv.writer(response)
    writer.writerow(["Student ID", "Semester ID", "Course ID", "Grade"])
    for sid, sem, cid, g in zip(student_ids, semester_ids, course_ids, grades):
        writer.writerow([sid, sem, cid, g])
    return response


# ---------------------------------------------------------------------------
# Submit grades prof — course list
# ---------------------------------------------------------------------------

def get_prof_courses_data(username, academic_year, semester_type, programme_type=None):
    """Return courses data for a professor."""
    working_year, _ = parse_academic_year(academic_year, semester_type)
    courses_query = selectors.get_instructor_courses(username, working_year, semester_type)

    student_ids_with_programme = None
    if programme_type:
        prog_list = selectors.get_programme_list(programme_type)
        if prog_list:
            student_ids_with_programme = Student.objects.filter(
                programme__in=prog_list
            ).values_list("id", flat=True)
            course_ids_with_programme = (
                course_registration.objects.filter(
                    course_id__in=courses_query.values_list("id", flat=True),
                    student_id__in=student_ids_with_programme,
                    session=academic_year,
                    semester_type=semester_type,
                )
                .values_list("course_id", flat=True)
                .distinct()
            )
            courses_query = courses_query.filter(id__in=course_ids_with_programme)

    courses_data = []
    for course in courses_query.values():
        course_regs = course_registration.objects.filter(
            course_id=course["id"],
            session=academic_year,
            semester_type=semester_type,
        )
        if programme_type and student_ids_with_programme is not None:
            course_regs = course_regs.filter(student_id__in=student_ids_with_programme)
        course["student_count"] = course_regs.count()
        course["has_students"] = course["student_count"] > 0
        course["programme_type"] = programme_type
        courses_data.append(course)

    return courses_data


# ---------------------------------------------------------------------------
# Download grades (prof)
# ---------------------------------------------------------------------------

def get_download_grades_data(username, academic_year, semester_type, programme_type=None):
    """Return courses with submitted grades for a professor."""
    working_year, _ = parse_academic_year(academic_year, semester_type)
    instructor_course_ids = selectors.get_instructor_course_ids(
        username, working_year, semester_type
    )

    grades_qs = Student_grades.objects.filter(
        academic_year=academic_year,
        semester_type=semester_type,
        course_id_id__in=instructor_course_ids.values_list("course_id_int", flat=True),
    )

    if programme_type:
        prog_list = selectors.get_programme_list(programme_type)
        if not prog_list:
            raise ValueError("Invalid programme_type. Must be 'UG' or 'PG'.")
        student_ids = Student.objects.filter(programme__in=prog_list).values_list("id", flat=True)
        grades_qs = grades_qs.filter(roll_no__in=student_ids)

    course_ids = grades_qs.values_list("course_id_id", flat=True).distinct()
    return list(Courses.objects.filter(id__in=course_ids).values())


# ---------------------------------------------------------------------------
# Generate course grade PDF (faculty)
# ---------------------------------------------------------------------------

def generate_course_grade_pdf(
    course_id, academic_year, semester_type, request_user, programme_type=None
):
    """Generate PDF with grade sheet for a course. Returns HttpResponse."""
    course_info = Courses.objects.get(id=course_id)
    working_year, _ = parse_academic_year(academic_year, semester_type)

    grades = Student_grades.objects.filter(
        course_id_id=course_id,
        academic_year=academic_year,
        semester_type=semester_type,
    )

    if programme_type:
        prog_list = selectors.get_programme_list(programme_type)
        if not prog_list:
            raise ValueError("Invalid programme_type. Must be 'UG' or 'PG'.")
        student_ids = Student.objects.filter(programme__in=prog_list).values_list("id", flat=True)
        grades = grades.filter(roll_no__in=student_ids)

    grades = grades.order_by("roll_no")

    ci = CourseInstructor.objects.filter(
        course_id_id=course_id,
        year=working_year,
        semester_type=semester_type,
        instructor_id_id=request_user.username,
    )
    if not ci.exists():
        return None, "Course not found."

    instructor = f"{request_user.first_name} {request_user.last_name}"
    grade_counts = {g: grades.filter(grade=g).count() for g in ALL_DISPLAY_GRADES}

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{course_info.code}_grades.pdf"'

    pdf_title = f"Course Grades - {course_info.code} - {course_info.name}"
    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=inch,
        rightMargin=inch,
        topMargin=2 * inch,
        bottomMargin=inch,
        title=pdf_title,
        author="PDPM IIITDM Jabalpur",
        subject=f"Course Grade Report - {course_info.code}",
        creator="Fusion Academic System",
    )

    elements = []
    styles = getSampleStyleSheet()

    field_label_style = ParagraphStyle(
        "FieldLabelStyle", parent=styles["Normal"],
        fontSize=12, textColor=colors.black, spaceAfter=5,
    )
    header_style = ParagraphStyle(
        "HeaderStyle", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=16,
        textColor=HexColor("#333333"), spaceAfter=20, alignment=1,
    )

    data = [["S.No.", "Roll Number", "Grade"]]
    for i, g in enumerate(grades, 1):
        data.append([i, g.roll_no, g.grade])
    tbl = Table(data, colWidths=[80, 300, 100])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E0E0E0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 14),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#F9F9F9"), colors.white]),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 12),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 20))

    grade_data1 = [["O", "A+", "A", "B+", "B", "C+", "C", "D+"]]
    grade_data1.append([grade_counts[g] for g in grade_data1[0]])
    grade_table1 = Table(grade_data1, colWidths=[60] * 8)
    grade_table1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E0E0E0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(grade_table1)
    elements.append(Spacer(1, 10))

    grade_data2 = [["D", "F", "S", "X", "CD"]]
    grade_data2.append([grade_counts[g] for g in grade_data2[0]])
    grade_table2 = Table(grade_data2, colWidths=[80] * 5)
    grade_table2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E0E0E0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(grade_table2)
    elements.append(Spacer(1, 40))

    verified_style = ParagraphStyle(
        "VerifiedStyle", parent=styles["Normal"],
        fontSize=13, textColor=HexColor("#333333"), spaceAfter=20,
    )
    elements.append(Paragraph(
        "I have carefully checked and verified the submitted grades. "
        "The grade distribution and submitted grades are correct. "
        "[Please mention any exception below.]",
        verified_style,
    ))

    def draw_page(canvas_obj, doc_obj):
        canvas_obj.setTitle(f"Grade Sheet - {course_info.code}")
        canvas_obj.saveState()
        width, height = letter
        p_title = Paragraph("Grade Sheet", header_style)
        w, h = p_title.wrap(doc_obj.width, doc_obj.topMargin)
        p_title.drawOn(canvas_obj, doc_obj.leftMargin, height - h)
        course_details = (
            f"L:{course_info.lecture_hours}, T:{course_info.tutorial_hours}, "
            f"P:{course_info.project_hours}, C:{course_info.credit}"
        )
        hdr_texts = [
            f"<b>Session:</b> {academic_year}",
            f"<b>Course Code:</b> {course_info.code}",
            f"<b>Course Name:</b> {course_info.name} ({course_details})",
            f"<b>Instructor:</b> {instructor}",
        ]
        y = height - h - header_style.spaceAfter
        for txt in hdr_texts:
            p = Paragraph(txt, field_label_style)
            w2, h2 = p.wrap(doc_obj.width, doc_obj.topMargin)
            p.drawOn(canvas_obj, doc_obj.leftMargin, y - h2)
            y -= (h2 + field_label_style.spaceAfter)
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.drawRightString(width - inch, 0.3 * inch, f"Page {doc_obj.page}")
        canvas_obj.setFont("Helvetica", 12)
        canvas_obj.drawString(inch, 0.75 * inch, "")
        canvas_obj.drawString(inch, 0.5 * inch, "Date")
        canvas_obj.drawString(width - 4 * inch, 0.75 * inch, "")
        canvas_obj.drawString(width - 4 * inch, 0.5 * inch, "Course Instructor's Signature")
        canvas_obj.restoreState()

    doc.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)
    return response, None


# ---------------------------------------------------------------------------
# Generate student result PDF
# ---------------------------------------------------------------------------

def generate_student_result_pdf(
    student_info, courses, spi, cpi, su, tu,
    semester_no, semester_type, semester_label="",
    is_transcript=False,
):
    """Generate a PDF for a student's result. Returns HttpResponse."""
    formatted_semester = format_semester_display(semester_no, semester_type, semester_label)
    buffer = BytesIO()

    doc_type = "Transcript" if is_transcript else "Student Result"
    pdf_title = (
        f"{doc_type} - "
        f"{student_info.get('name', student_info.get('rollNumber', 'Student'))} - "
        f"{formatted_semester}"
    )

    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=0.5 * inch, leftMargin=0.5 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        title=pdf_title,
        author="PDPM IIITDM Jabalpur",
        subject=f"{doc_type} Report - {formatted_semester}",
        creator="Fusion Academic System",
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"],
        fontSize=14, spaceAfter=2, leading=18,
        alignment=1, fontName="Times-Bold",
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=11, spaceAfter=6,
        alignment=1, fontName="Times-Roman",
    )

    # Header with optional logo
    try:
        logo_path = os.path.join(settings.MEDIA_ROOT, "logo2.jpg")
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=0.8 * inch, height=0.8 * inch)
            title_para = Paragraph(
                "PDPM Indian Institute of Information Technology,", title_style
            )
            subtitle1_para = Paragraph("Design & Manufacturing, Jabalpur", title_style)
            subtitle2_para = Paragraph(
                "(An Institute of National Importance under MoE, Govt. of India)",
                subtitle_style,
            )
            subtitle3_para = Paragraph(
                "<b><u>Semester Grade Report / Marksheet</u></b>", subtitle_style
            )
            header_table_data = [
                [logo, [title_para, subtitle1_para, subtitle2_para, subtitle3_para]]
            ]
            header_table = Table(header_table_data, colWidths=[1 * inch, 6 * inch])
            header_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(header_table)
        else:
            _add_text_header(story, title_style, subtitle_style)
    except Exception:
        _add_text_header(story, title_style, subtitle_style)

    story.append(Spacer(1, 12))

    # Student info table
    cell_style = ParagraphStyle(
        "CellStyle", parent=styles["Normal"],
        fontSize=10, fontName="Times-Roman", wordWrap="CJK",
    )

    roll_no_key = student_info.get("rollNumber", student_info.get("roll_number", "N/A"))
    branch_key = student_info.get("branch", student_info.get("department", "N/A"))
    acad_year_key = student_info.get("academicYear", student_info.get("academic_year", "N/A"))

    student_data = [
        [
            Paragraph("Name of Student:", cell_style),
            Paragraph(student_info.get("name", "N/A"), cell_style),
            Paragraph("Roll No.:", cell_style),
            Paragraph(roll_no_key, cell_style),
        ],
        [
            Paragraph("Programme:", cell_style),
            Paragraph(student_info.get("programme", "N/A"), cell_style),
            Paragraph("Branch:", cell_style),
            Paragraph(branch_key, cell_style),
        ],
        [
            Paragraph("Semester:", cell_style),
            Paragraph(formatted_semester, cell_style),
            Paragraph("Academic Year:", cell_style),
            Paragraph(acad_year_key, cell_style),
        ],
    ]

    student_table = Table(
        student_data, colWidths=[1.14 * inch, 3.56 * inch, 1.3 * inch, 1.0 * inch]
    )
    student_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(student_table)
    story.append(Spacer(1, 12))

    # Course table
    headers = ["S. No.", "Course Code", "Course Title", "Credits", "Grade", "Grade Points"]
    course_data = [headers]
    for i, course in enumerate(courses, 1):
        course_data.append([
            str(i),
            course.get("coursecode", ""),
            course.get("coursename", ""),
            str(course.get("credits", "")),
            course.get("grade", ""),
            str(course.get("points", "")),
        ])

    course_table = Table(
        course_data,
        colWidths=[0.5 * inch, 1 * inch, 3.2 * inch, 0.7 * inch, 0.6 * inch, 1 * inch],
    )
    course_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Times-Roman"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(course_table)
    story.append(Spacer(1, 15))

    # Summary
    summary_data = [
        ["Total Credits Registered:", str(tu), "Semester Credits Earned:", str(su)],
        ["SPI:", f"{float(spi):.1f}", "CPI:", f"{float(cpi):.1f}"],
    ]
    summary_table = Table(
        summary_data, colWidths=[2.2 * inch, 0.8 * inch, 2.2 * inch, 0.8 * inch]
    )
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 25))

    # Footer
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontSize=8, alignment=1, fontName="Times-Italic",
    )
    current_date = datetime.now().strftime("%d/%m/%Y")
    story.append(Paragraph(
        f"This is a computer-generated document. Generated on {current_date}",
        footer_style,
    ))

    doc.build(story)

    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type="application/pdf")
    semester_suffix = formatted_semester.replace(" ", "_").replace(":", "").lower()
    prefix = "transcript_" if is_transcript else "result_"
    filename = (
        f"{prefix}"
        f"{student_info.get('rollNumber', student_info.get('roll_number', 'student'))}_"
        f"{semester_suffix}.pdf"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Content-Length"] = len(pdf_data)
    return response


def _add_text_header(story, title_style, subtitle_style):
    story.append(Paragraph("PDPM Indian Institute of Information Technology,", title_style))
    story.append(Paragraph("Design & Manufacturing, Jabalpur", title_style))
    story.append(Paragraph(
        "(An Institute of National Importance under MoE, Govt. of India)", subtitle_style
    ))
    story.append(Paragraph(
        "<b><u>Semester Grade Report / Marksheet</u></b>", subtitle_style
    ))


# ---------------------------------------------------------------------------
# Fetch student result data (for CheckResultView)
# ---------------------------------------------------------------------------

def get_student_result_data(roll_number, semester_no, semester_type):
    """
    Return result data dict or (None, error_dict).
    """
    try:
        student = selectors.get_student_by_roll(roll_number)
    except Student.DoesNotExist:
        return None, {"success": False, "message": "Student record not found."}

    ann = selectors.get_result_announcement(student.batch_id, semester_no)
    if not ann or not ann.announced:
        return None, {"success": False, "message": "Results not announced yet."}

    grades_info = selectors.get_student_grades_for_student(
        roll_number, semester_no, semester_type
    )

    academic_year = None
    if grades_info.exists():
        academic_year = grades_info.first().academic_year

    spi, su, _ = calculate_spi_for_student(student, semester_no, semester_type)
    cpi, tu, _ = calculate_cpi_for_student(student, semester_no, semester_type)

    student_info = _build_student_info(student, academic_year)

    courses = [
        {
            "coursecode": g.course_id.code,
            "courseid": g.course_id.id,
            "coursename": g.course_id.name,
            "credits": g.course_id.credit,
            "grade": g.grade,
            "points": compute_grade_points(g.grade),
        }
        for g in grades_info
    ]

    return {
        "success": True,
        "student_info": student_info,
        "courses": courses,
        "spi": spi,
        "cpi": cpi,
        "su": su,
        "tu": tu,
    }, None


# ---------------------------------------------------------------------------
# Student result PDF (from DB, used by GenerateStudentResultPDFAPI)
# ---------------------------------------------------------------------------

def generate_student_result_pdf_from_db(roll_number, semester_no, semester_type, request_data):
    """Fetch data from DB and generate PDF. Returns HttpResponse or error dict."""
    result_data, error = get_student_result_data(roll_number, semester_no, semester_type)
    if error:
        return None, error

    semester_label = request_data.get("semester_label", "")
    is_transcript = (
        request_data.get("is_transcript", False)
        or request_data.get("document_type") == "transcript"
    )

    return generate_student_result_pdf(
        student_info=result_data["student_info"],
        courses=result_data["courses"],
        spi=result_data["spi"],
        cpi=result_data["cpi"],
        su=result_data["su"],
        tu=result_data["tu"],
        semester_no=semester_no,
        semester_type=semester_type,
        semester_label=semester_label,
        is_transcript=is_transcript,
    ), None


# ---------------------------------------------------------------------------
# Validate dean CSV
# ---------------------------------------------------------------------------

def validate_dean_csv(course_id, academic_year, csv_file):
    """Compare CSV grades against DB. Returns mismatches list or error."""
    decoded_file = csv_file.read().decode("utf-8").splitlines()
    reader = csv.DictReader(decoded_file)
    required_columns = ["roll_no", "grade", "remarks"]
    if not all(col in (reader.fieldnames or []) for col in required_columns):
        return None, "CSV file must contain the following columns: roll_no, grade, remarks."

    mismatches = []
    for row in reader:
        roll_no = row["roll_no"]
        grade_val = row["grade"]
        remarks = row["remarks"]

        try:
            student = Student.objects.get(id_id=roll_no)
            semester = student.curr_semester_no
            batch = student.batch

            student_grade = Student_grades.objects.get(
                roll_no=roll_no,
                course_id_id=course_id,
                year=academic_year,
                batch=batch,
            )

            if student_grade.grade != grade_val:
                mismatches.append({
                    "roll_no": roll_no,
                    "csv_grade": grade_val,
                    "db_grade": student_grade.grade,
                    "remarks": remarks,
                    "batch": batch,
                    "semester": semester,
                    "course_id": course_id,
                })
        except (Student.DoesNotExist, Student_grades.DoesNotExist):
            return None, f"Student or grade record not found for roll number: {roll_no}"

    return mismatches, None


# ---------------------------------------------------------------------------
# Preview grades CSV
# ---------------------------------------------------------------------------

def preview_grades_csv(course_id, academic_year, semester_type, csv_file, programme_type=None):
    """Parse CSV and return preview with registration status. Returns (preview_rows, error)."""
    try:
        working_year = int(academic_year.split("-")[0])
    except Exception:
        return None, "Invalid academic_year format. Expected format like 2023-24."

    registrations = course_registration.objects.filter(
        course_id=course_id, session=academic_year, semester_type=semester_type,
    )

    if programme_type:
        prog_list = selectors.get_programme_list(programme_type)
        if not prog_list:
            return None, "Invalid programme_type. Must be 'UG' or 'PG'."
        student_ids = Student.objects.filter(programme__in=prog_list).values_list("id", flat=True)
        registrations = registrations.filter(student_id__in=student_ids)

    registered_rollnos = set()
    for reg in registrations.select_related("student_id"):
        if hasattr(reg.student_id, "id_id"):
            registered_rollnos.add(reg.student_id.id_id)
        else:
            registered_rollnos.add(str(reg.student_id_id))

    try:
        decoded_file = csv_file.read().decode("utf-8")
        io_string = StringIO(decoded_file)
        reader = csv.DictReader(io_string)
    except Exception as e:
        return None, f"Error reading CSV file: {str(e)}"

    required_columns = ["roll_no", "name", "grade", "remarks", "semester"]
    if not all(col in (reader.fieldnames or []) for col in required_columns):
        return None, f"CSV file must contain the following columns: {', '.join(required_columns)}"

    preview_rows = []
    for row in reader:
        roll_no = row["roll_no"]
        preview_rows.append({
            "roll_no": roll_no,
            "name": row["name"],
            "branch": row.get("branch", ""),
            "grades": row["grade"],
            "remarks": row["remarks"],
            "semester": row["semester"],
            "is_registered": roll_no in registered_rollnos,
        })

    return preview_rows, None


# ---------------------------------------------------------------------------
# Result announcements
# ---------------------------------------------------------------------------

def get_announcements_data(role):
    """Return announcements and batch options."""
    announcements = selectors.get_result_announcements()
    ann_data = []
    for ann in announcements:
        batch = ann.batch
        batch_label = f"{batch.name} - {batch.discipline.acronym} {batch.year}"
        ann_data.append({
            "id": ann.id,
            "batch": {
                "id": batch.id,
                "name": batch.name,
                "discipline": batch.discipline.acronym,
                "year": batch.year,
                "label": batch_label,
            },
            "semester": ann.semester,
            "announced": ann.announced,
            "created_at": ann.created_at,
        })

    batch_objs = selectors.get_running_batches()
    batch_options = []
    for b in batch_objs:
        label = f"{b.name} - {b.discipline.acronym} {b.year}"
        batch_options.append({"id": b.id, "label": label})

    return {"announcements": ann_data, "batches": batch_options}


def update_announcement(announcement_id, announced):
    ann = ResultAnnouncement.objects.get(id=announcement_id)
    ann.announced = announced
    ann.save()
    return True


def create_announcement(batch_id, semester):
    batch_obj = selectors.get_batch_by_id(batch_id)
    if not batch_obj:
        return None, "Batch not found."

    ann, created = ResultAnnouncement.objects.get_or_create(
        batch=batch_obj, semester=semester, defaults={"announced": False}
    )
    batch_label = f"{batch_obj.name} - {batch_obj.discipline.acronym} {batch_obj.year}"
    data = {
        "id": ann.id,
        "batch": {"id": batch_obj.id, "label": batch_label},
        "semester": ann.semester,
        "announced": ann.announced,
        "created_at": ann.created_at,
    }
    return data, created


# ---------------------------------------------------------------------------
# Student semester list
# ---------------------------------------------------------------------------

def get_student_semester_list(roll_number):
    qs = selectors.get_student_semester_list(roll_number)
    unique = OrderedDict()
    for sem_no, sem_type in qs:
        label = make_semester_label(sem_no, sem_type or "")
        unique[(sem_no, sem_type)] = label
    return [
        {"semester_no": no, "semester_type": typ, "label": lbl}
        for (no, typ), lbl in unique.items()
    ]


# ---------------------------------------------------------------------------
# Grade status
# ---------------------------------------------------------------------------

def get_grade_status(academic_year, semester_type):
    """Return grade status list for all courses in given year/semester."""
    working_year, session = parse_academic_year(academic_year, semester_type)

    course_ids = (
        course_registration.objects.filter(
            session=academic_year, semester_type=semester_type
        )
        .values_list("course_id", flat=True)
        .distinct()
    )
    courses = Courses.objects.filter(id__in=course_ids).order_by("code")

    instructors_map = {}
    for inst in selectors.get_course_instructors_bulk(course_ids, working_year, semester_type):
        instructors_map[inst.course_id_id] = inst

    instructor_ids = [inst.instructor_id_id for inst in instructors_map.values()]
    users_map = selectors.get_users_by_usernames(instructor_ids) if instructor_ids else {}

    submitted_courses = selectors.get_submitted_course_ids(course_ids, academic_year, semester_type)
    verified_courses = selectors.get_verified_course_ids(course_ids, academic_year, semester_type)

    auth_records_map = {}
    for auth in selectors.get_authentication_records_bulk(course_ids, working_year):
        auth_records_map[auth.course_id_id] = auth

    grade_status_list = []
    for course in courses:
        instructor = instructors_map.get(course.id)
        professor_name = "Not Assigned"
        if instructor:
            professor_name = users_map.get(
                instructor.instructor_id_id, instructor.instructor_id_id
            )

        submitted = "Submitted" if course.id in submitted_courses else "Not Submitted"
        verified = "Verified" if course.id in verified_courses else "Not Verified"

        validated = "Not Validated"
        if course.id in verified_courses:
            auth_record = auth_records_map.get(course.id)
            if (
                auth_record
                and auth_record.authenticator_1
                and auth_record.authenticator_2
                and auth_record.authenticator_3
            ):
                validated = "Validated"

        grade_status_list.append({
            "course_code": course.code,
            "course_name": course.name,
            "course_id": course.id,
            "professor_name": professor_name,
            "submitted": submitted,
            "verified": verified,
            "validated": validated,
            "credits": course.credit,
            "version": course.version,
        })

    return grade_status_list


# ---------------------------------------------------------------------------
# Grade summary (raw SQL)
# ---------------------------------------------------------------------------

def get_grade_summary(academic_year, semester_type):
    """Return grade summary using raw SQL."""
    query = """
        SELECT
            ROW_NUMBER() OVER (ORDER BY pc.code) as sno,
            pc.code as course_code,
            pc.name as course_name,
            STRING_AGG(DISTINCT TRIM(CONCAT(u.first_name, ' ', u.last_name)), ', ') as course_instructor,
            COUNT(CASE WHEN sg.grade = 'O' THEN 1 END) as grade_o,
            COUNT(CASE WHEN sg.grade = 'A+' THEN 1 END) as grade_a_plus,
            COUNT(CASE WHEN sg.grade = 'A' THEN 1 END) as grade_a,
            COUNT(CASE WHEN sg.grade = 'B+' THEN 1 END) as grade_b_plus,
            COUNT(CASE WHEN sg.grade = 'B' THEN 1 END) as grade_b,
            COUNT(CASE WHEN sg.grade = 'C+' THEN 1 END) as grade_c_plus,
            COUNT(CASE WHEN sg.grade = 'C' THEN 1 END) as grade_c,
            COUNT(CASE WHEN sg.grade = 'D+' THEN 1 END) as grade_d_plus,
            COUNT(CASE WHEN sg.grade = 'D' THEN 1 END) as grade_d,
            COUNT(CASE WHEN sg.grade = 'F' THEN 1 END) as grade_f,
            COUNT(CASE WHEN sg.grade = 'CD' THEN 1 END) as grade_cd,
            COUNT(CASE WHEN sg.grade = 'S' THEN 1 END) as grade_s,
            COUNT(CASE WHEN sg.grade = 'X' THEN 1 END) as grade_x,
            COUNT(sg.id) as total_students
        FROM
            online_cms_student_grades sg
            INNER JOIN programme_curriculum_course pc ON sg.course_id_id = pc.id
            LEFT JOIN programme_curriculum_courseinstructor ci ON (
                ci.course_id_id = pc.id
                AND ci.year = sg.year
            )
            LEFT JOIN auth_user u ON ci.instructor_id_id = u.username
        WHERE
            sg.academic_year = %s
            AND sg.semester_type = %s
            AND sg.grade IS NOT NULL
            AND sg.grade <> ''
        GROUP BY
            pc.code, pc.name
        HAVING
            COUNT(sg.id) > 0
        ORDER BY
            pc.code
    """
    with connection.cursor() as cursor:
        cursor.execute(query, [academic_year, semester_type])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]
